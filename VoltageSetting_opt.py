'''
################################################################
# 脚本用来实现电压插值.
#
# 输入文件：
# - json文件
# - energyfactor.csv
#
# 输出文件：
# - 插值电压列表.csv
# - 修改后的json
# - 获得目标文件的温度分布情况
# - 获得电压分布曲线图
#
# Author: Jerome. Date: 20240626
#
#################################################################
#
# 1st 修改
# 修改目的：
# - autotune 的 json 复制后修改后缀为 「***_opt_0.json」在判定 optimise 文件夹不存在时一起处理
# - 电压插值法优化处理完毕后，在 optimise 文件夹中保留最新的 「***_opt_N.json」
#   并将最新的 json 文件复制并移动到上级目录 「\InUse」中，重命名为 「***_for_calib_Tune.json」替换原来的文件
#
# 修改完成时间：20240730
#
################################################################

'''

import time
import os
import sys
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
import json
import shutil
from Function.envi import ENVI
from Function.UNSFunction import *
from datetime import datetime
import json
import csv
import glob
import openpyxl
from datetime import datetime

def get_json_data(json_path):
    # 读取 json 文件内容
    with open(json_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data

def get_csv_data(csv_path):
    # 打开并读取 CSV 文件
    with open(csv_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        # 将内容保存为矩阵
        matrix = [row for row in reader]
        # 打印矩阵内容
        # for row in matrix:
        #     print(row)

    return matrix

class StructVar:
    # 建立存储电压的类
    def __init__(self):
        self.v1 = []
        self.v2 = []
        self.v3 = []
        self.v4 = []


def fitting_voltage(json_path, csv_path, MEMS_params):

    MEMS_data_array = MEMS_params[0]
    MEMS_ID = MEMS_params[1]

    energyfactor_index = next((i for i, path in enumerate(csv_path) if 'energyFactor.csv' in path), None)
    # 获取标定的 cwl 矩阵 来自 energyFactor.csv 文件
    csv_data = get_csv_data(csv_path[energyfactor_index])
    cwl_cal = []
    cwl_warning = []
    differ_warning = []
    cwl_indx = []
    for i in range(10):
        data = int(csv_data[i + 1][0])
        cwl_cal.append((data))
        if abs(int(csv_data[i + 1][2])) >= 4:
            print(f'-> Info: band_{i+1} cwl is exceeds grade 4 nm.')
            print(f'-> Info: CWL is {data}, differ is {int(csv_data[i + 1][2])} nm.')
            cwl_warning.append(data)
            differ_warning.append(int(csv_data[i + 1][2]))
            cwl_indx.append(i+1)

    correction_factor = [cwl_indx, cwl_warning, differ_warning]

    ################################
    # 获取待重构的json文件 来自 energyFactor.csv 文件
    json_data = get_json_data(json_path)
    band_size = len(json_data['NIR']['MEMS']['Modes'])
    json_voltage_original = []

    json_voltage = StructVar()
    json_voltage_fit = StructVar()

    for i in range(band_size):
        json_voltage_original.append(json_data['NIR']['MEMS']['Modes'][i]['Voltages'])
        json_voltage.v1.append(json_voltage_original[i][0])
        json_voltage.v2.append(json_voltage_original[i][1])
        json_voltage.v3.append(json_voltage_original[i][2])
        json_voltage.v4.append(json_voltage_original[i][3])

    # 判断此 MEMS 是 1st order 还是 2nd order
    json_voltage_interval = cal_json_voltage_model(json_voltage.v1)

    if len(json_voltage_interval) > 1:
        print(f'-> Info: MEMS voltages is multimodal!')
    else:
        print(f'-> Info: MEMS voltage is unimodal!')

    # plot修改之前的json电压
    # 循环访问原始电压中的所有变量，进行运算后保存到新的数组中
    x_array_original = cwl_cal
    for key in json_voltage.__dict__:
        value = getattr(json_voltage, key)
        f_linear = interp1d(x_array_original, value)

        plt.figure(3, figsize=(10, 6))
        plt.plot(x_array_original, getattr(json_voltage, key), marker='o', color='black', label='Before Voltage Treatment')
        plt.grid(True)

    #TODO: 基于获得的 json_voltage 及 energyfactor 表格，开始对电压趋势的判断
    count_num = 1
    exemptions_array = []
    for i in range(len(cwl_indx)):
        # 用于记录使用 MEMS data array 推导出来的索引，用于豁免电压差值
        # 首先处理两端超界限的条件
        if (cwl_indx[i] == 1 and differ_warning[i] <= -5) or (cwl_indx[i] == 10 and differ_warning[i] >= 5):
            e_indx = cwl_indx[i]
            exemptions_array.append(e_indx)
            # 边界电压判断存在问题
            [json_voltage, correction_factor] = boundary_voltage_calculation(i, MEMS_ID, MEMS_data_array, correction_factor, json_voltage)
            # 更新此时的 cwl_cal
            # json_voltage 及 correction_factor 两组参数已经更新
            for j in range(len(correction_factor[0])):
                cwl_cal[cwl_indx[j]-1] = correction_factor[1][j]
            count_num = count_num + 1

        elif any(cwl_indx[i]-1 == start or cwl_indx[i]-1 == end for start, end in json_voltage_interval):
            e_indx = cwl_indx[i]
            exemptions_array.append(e_indx)
            # 边界电压判断存在问题
            [json_voltage, correction_factor] = boundary_voltage_calculation(i, MEMS_ID, MEMS_data_array,
                                                                             correction_factor, json_voltage)
            # 更新此时的 cwl_cal
            # json_voltage 及 correction_factor 两组参数已经更新
            for j in range(len(correction_factor[0])):
                cwl_cal[cwl_indx[j]-1] = correction_factor[1][j]

    cwl = [713, 736, 759, 782, 805, 828, 851, 874, 897, 920]

    # 计算差值点数
    voltage_number = cwl_cal[-1] - cwl_cal[0] + 1
    x_array_original = cwl_cal

    # 创建一组新的x值，用于插值计算
    x_array_new = np.linspace(cwl_cal[0], cwl_cal[-1], num=voltage_number)
    cwl_array_fit = np.linspace(cwl_cal[0], cwl_cal[-1], num=voltage_number)

    print(f'\n-> Info: cwl_cal array is {x_array_original}')
    dataBeSave = np.zeros([5, voltage_number])
    dataBeSave[0, :] = cwl_array_fit
    count = 1
    # 循环访问原始电压中的所有变量，进行运算后保存到新的数组中
    for key in json_voltage.__dict__:
        value = getattr(json_voltage, key)
        f_linear = interp1d(x_array_original, value)
        # 进行运算，这里是加10
        setattr(json_voltage_fit, key, f_linear(x_array_new))

        dataBeSave[count, :] = getattr(json_voltage_fit, key)
        count = count + 1

        plt.figure(3, figsize=(10, 6))
        plt.title('Boundary voltage trends')
        plt.plot(x_array_original, getattr(json_voltage, key), marker='o', color='red', label='After Voltage Treatment')
        plt.plot(x_array_new, getattr(json_voltage_fit, key), color='blue', linewidth=2, label='After Voltage Treatment Fitting')
        plt.xlabel('Wavelength / nm')
        plt.ylabel('Voltage / V')
        plt.grid(True)

    dataBeSave = dataBeSave.T
    return dataBeSave, correction_factor, exemptions_array

def save_xlsx(folder_save_Path, dataBeSave):
    dataHead = ['cwl_cal', 'V1', 'V2', 'V3', 'V4']
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    array_dataBeSave = np.array(dataBeSave)
    # 确定另存的hyperspectral数据都是有效的，对于2nd order数据进行mask处理
    [increasing_to_decreasing, decreasing_to_increasing] = find_monotonic_changes(dataBeSave[:, 1])
    for i in range(len(decreasing_to_increasing)-1):
        for j in range(4):
            dataBeSave[int(min(decreasing_to_increasing[0])+i+1), j+1] = 0

    # 将字符串列表写入 Excel 的第一行
    for index, value in enumerate(dataHead, start=1):
        worksheet.cell(row=1, column=index, value=value)

    # 将二维数字列表写入 Excel 从第二行开始的不同行
    for row_index, row_data in enumerate(dataBeSave, start=2):
        for col_index, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    # 生成保存 Excel 文件的完整路径
    check_file_number = glob.glob(f"{folder_save_Path}/*.xlsx")
    file_number = len(check_file_number)

    excel_file = folder_save_Path + f'\\voltage_fitting_opt_{file_number+1}.xlsx'
    print(f'-> Info: new_file_name is {excel_file}')

    # excel_file = folder_save_Path + r'\voltage_fitting.xlsx'
    # 保存工作簿为 Excel 文件
    try:
        workbook.save(excel_file)
        print(f"-> Info: Data saved to {excel_file}")
    except:
        print('-> Info: Please close csv file and try again.')



def save_fitting_json(file_path, dataBeSave, json_path, json_index, exemptions_array):
    current_time = datetime.now()
    formatted_time = current_time.strftime("%Y%m%d%H%M")
    # 假设 'original.json' 是原始文件名，'modified.json' 是新文件名
    original_file_name = json_path[json_index]
    # 提取文件名和扩展名
    file_name_with_extension = os.path.basename(json_path[json_index])
    # 去掉.json后缀
    file_name_without_extension = os.path.splitext(file_name_with_extension)[0]
    # 创建新json的存放文件夹
    foldername = '\WithMEMS\SystemResponse\InUse\optimise'
    folder_save_Path = CreateFolder(file_path, foldername)

    check_file_number = glob.glob(f"{folder_save_Path}/*.json")
    file_number = len(check_file_number)
    if file_number == 0:
        # 将初次 autotune 的 json file 复制并重命名到 optimise 文件夹下
        primary_file_name = folder_save_Path + f'\\{file_name_without_extension}_opt_0.json'
        shutil.copy(original_file_name, primary_file_name)
        print(f'-> Info: This is first time to deal with autotune calib json file.')

        # 建立当下将要生成的 json file 的文件名
        new_file_name = folder_save_Path + f'\\{file_name_without_extension}_opt_{file_number+1}.json'

    else:
        file_name_without_extension = re.sub(r'\d+$', '', file_name_without_extension)
        new_file_name = folder_save_Path + f'\\{file_name_without_extension}{file_number}.json'
    print(f'-> Info: new_file_name is {new_file_name}')

    try:
        # 加载原始JSON数据
        json_data = get_json_data(original_file_name)

        for i in range(len(correction_factor[0])):
            print(f'\n-> Info: Now fix the parameters of band {correction_factor[0][i]} ')
            # 原来 json 文件中的电压为
            indx = correction_factor[0][i] - 1
            A = json_data['NIR']['MEMS']['Modes'][indx]['Voltages']
            print(f'-> Info: Original voltage array is {A}')

            #TODO:计算正确的 cwl
            # 需要注意绕过豁免索引矩阵中的内容
            if correction_factor[0][i] not in exemptions_array:
                cwl_corrected = correction_factor[1][i] + correction_factor[2][i]
            else:
                cwl_corrected = correction_factor[1][i]
            # 找到 dataBeSave 中正确 cwl 对应的电压
            indexes_of_value = np.where(dataBeSave[:, 0] == cwl_corrected)
            print(f'-> Info: Indx is {indexes_of_value[0]}, corrected cwl is {cwl_corrected}')
            voltage_corrected = [float(dataBeSave[indexes_of_value, 1]), float(dataBeSave[indexes_of_value, 2]), float(dataBeSave[indexes_of_value, 3]), float(dataBeSave[indexes_of_value, 4])]
            print(f'-> Info: Corrected voltage is {voltage_corrected}')

            # 对原来的电压进行修改
            json_data['NIR']['MEMS']['Modes'][indx]['Voltages'] = voltage_corrected

            # 将修改后的数据保存到新的JSON文件
            with open(new_file_name, 'w') as file:
                json.dump(json_data, file, indent=4)

        print(f'\n-> Info: Successfully saved the optimise file!')

        # 将新生成的 opt 的 json 复制并替换上级目录中的 最终格式的json
        shutil.copy(new_file_name, original_file_name)

        print(f'\n-> Info: A file with the standard format name was generated as required!')


        return folder_save_Path
    except FileNotFoundError:
        print(f"文件 '{original_file_name}' 未找到。请确保文件路径正确。")
    except json.JSONDecodeError:
        print(f"文件 '{original_file_name}' 不是有效的JSON格式。")
    except Exception as e:
        print(f"保存文件时出现错误: {e}")


def get_json_folder(file_path):
    json_mid_path = r'\WithMEMS\SystemResponse\InUse'
    # 找到 Solomon 的文件层
    camera_layer = os.path.basename(os.path.normpath(file_path))
    # 使用split方法按下划线分割字符串，并取第一个部分
    camera_id = camera_layer.split('_')[0]
    camera_id_modified = camera_id[0].lower() + camera_id[1:]
    # 建立json文件的path
    json_third_path = file_path + json_mid_path

    # 判断是否已经建立了 optimise 文件夹
    opt_folder_path = json_third_path + r'\optimise'
    file_exists = check_file_exists(opt_folder_path)

    if file_exists and len(os.listdir(opt_folder_path)) > 0:
        # 如果已经存在了 optimise 文件夹，则基于已经优化过的 json 文件进一步优化
        json_path = glob.glob(f"{opt_folder_path}/*.json")
        check_file_number = len(json_path)
        json_file_pre = opt_folder_path + f'\\{camera_id_modified}_for_calib_Tune_opt_{check_file_number}.json'

    else:
        # 如果不存在 optimise 文件夹，则根据原始 tune 的 json 文件进行电压拟合
        json_path = glob.glob(f"{json_third_path}/*.json")
        json_file_pre = file_path + json_mid_path + f'\\{camera_id_modified}_for_calib_Tune.json'

    try:
        json_index = next((i for i, path in enumerate(json_path) if f'{json_file_pre}' in path), None)

    except:
        print('\n---------------------------------------------------------------')
        print('-> Error: Documentation missing!\n')
        json_path = []
        json_index = []

    return json_path, json_index


def get_csv_folder(file_path):
    # 提示用户输入
    print('-> Info: Choose which MEMS wanted to modified "1" is VIS, "2" is NIR')
    user_input = input("-> Info: Please input model ： ")
    # 判断输入是否为数字
    if user_input.isdigit():
        if int(user_input) == 1:
            csv_mid_path = r'\WithMEMS\SystemResponse\VIS_InUse'
            model = 'VIS'
        elif int(user_input) == 2:
            csv_mid_path = r'\WithMEMS\SystemResponse\NIR_InUse'
            model = 'NIR'
        else:
            print("-> Error: Please follow the rules to enter !")
    else:
        print("-> Error: Please input digital。")

    try:
        csv_path = glob.glob(f"{file_path + csv_mid_path}/*.csv")
        directory_path = file_path + csv_mid_path
        read_temp(directory_path)
    except:
        print('\n---------------------------------------------------------------')
        print('-> Error: Documentation missing!\n')
        csv_path = []

    return csv_path, model

def read_hdr_file(file_path):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            # print(len(content))
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {e}")


# 这是一个读取HDR文件并计算行数的示例代码
def count_lines_in_hdr_file(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            return len(lines)
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {e}")
        return 0

# 这是一个读取HDR文件，找到包含特定字符串的行并打印的示例代码
def find_temperature_line(file_path):
    try:
        with open(file_path, 'r') as file:
            for line in file:
                if 'temperature =' in line:
                    wanted_line = line.split()
                    # print(line.strip())
                    break

            return wanted_line
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {e}")

# 定义一个函数来列出目录下的所有文件
def list_files_in_directory(directory_path):
    # 获取目录下的所有文件和文件夹名称
    all_entries = os.listdir(directory_path)
    # 过滤出文件名称，忽略子目录
    file_names = [f for f in all_entries if os.path.isfile(os.path.join(directory_path, f))]
    return file_names

def get_wave(folder):
    # 使用正则表达式找到'state'后面紧跟着的数字
    match = re.search(r'state(\d+)', folder)

    # 如果找到匹配项，则提取数字
    if match:
        number = match.group(1)
        # print(f"'state'后面紧跟着的数字是: {number}")
        number = float(number)/10
    else:
        print("没有找到匹配的数字。")

    return number

def read_temp(directory_path):
    folder_list = os.listdir(directory_path)
    temp_box = []
    wave_box = []

    for folder in folder_list:
        # 找到 state 的文件
        if folder[0:5] == 'state':
            wave = get_wave(folder)
            file_path = os.path.join(directory_path, folder)
            # 过滤出.hdr文件
            all_entries = os.listdir(file_path)
            hdr_files = [f for f in all_entries if f.endswith('.hdr')]
            try:
                hdr_file_path = os.path.join(file_path,hdr_files[0])
                read_hdr_file(hdr_file_path)
                # 调用函数读取HDR文件并找到包含'temperature ='的行
                line_str = find_temperature_line(hdr_file_path)
                temp_data = round(float(line_str[2]), 2)
                temp_box.append(temp_data)
                wave_box.append(wave)
            except:
                print(f'{folder} 缺失了 hdr 文件')

    plt.figure(1, figsize=(10, 6))
    plt.title('Temperature')
    plt.plot(wave_box, temp_box, linewidth=2), plt.grid(True)


# 定义一个函数来检查文件是否存在
def check_file_exists(file_path):
    return os.path.exists(file_path)


def get_MEMS_data(json_path, model):
    json_data = get_json_data(json_path)
    MEMS_ID = json_data[f'{model}']['MEMS']['ID']

    # 加载 config 文件
    filepath = os.getcwd() + '\config.json'
    data = get_json_data(filepath)
    MEMS_ID_Path_1st = data['MEMS_data_folder_path']

    MEMS_ID_Path_2nd = f'\\{MEMS_ID}' + r'\Spectral analysis'
    MEMS_voltage_csv = glob.glob(f"{MEMS_ID_Path_1st + MEMS_ID_Path_2nd}/*.csv")
    if len(MEMS_voltage_csv) == 1:
        MEMS_data = get_csv_data(MEMS_voltage_csv[0])
    else:
        print('\n-> Info: Some error in folder!')
        if len(MEMS_voltage_csv) > 1:
            print('-> Info: Please enter which file is the avaliable one:')
            for i in range(len(MEMS_voltage_csv)):
                print(f'File {i} is {MEMS_voltage_csv[i]}')
            user_input = input("-> Info: Please input file No. ： ")
            # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! 判断输入是否为数字
            if user_input.isdigit():
                if int(user_input) > len(MEMS_voltage_csv):
                    print("-> Error: File number is over !")
                elif int(user_input) <= len(MEMS_voltage_csv):
                    MEMS_voltage_csv_indx = int(user_input)
                    MEMS_data = get_csv_data(MEMS_voltage_csv[MEMS_voltage_csv_indx])
                else:
                    print("-> Error: Please follow the rules to enter !")
            else:
                print("-> Error: Please input digital。")
        else:
            print(f'-> Info: File show as {MEMS_voltage_csv}')

    return MEMS_data, MEMS_ID


# 定义一个函数来找到数组中最接近目标值的索引
def find_closest(arr, target):
    # 将列表转换为numpy数组
    arr = np.array(arr)
    # 找到最接近值的索引
    index = (np.abs(arr - target)).argmin()
    return index


# 找到数组中单调递增的索引范围
def find_monotonic_intervals(arr, threshold):
    intervals = []
    start_idx = 0

    for i in range(1, len(arr)):
        if abs(arr[i] - arr[i-1]) > threshold:
            intervals.append((start_idx, i-1))
            start_idx = i

    intervals.append((start_idx, len(arr)-1))  # Add the last interval

    return intervals

def find_monotonic_intervals_without_threshold(arr):
    """
    查找数组中的单调区间，并返回每个区间的起始和结束索引。

    Args:
        arr (list): 输入的数组。

    Returns:
        list: 包含单调区间的元组列表，每个元组包含起始和结束索引。
    """
    intervals = []
    start = 0

    for i in range(1, len(arr)):
        if arr[i] < arr[i - 1] or arr[i] > arr[i - 1]:
            # 当前元素与前一个元素不相等，表示一个区间结束
            intervals.append((start, i - 1))
            start = i

    # 添加最后一个区间
    intervals.append((start, len(arr) - 1))

    return intervals


# 查找包含数字的范围的函数
def find_voltage_indx_range_in_LUT(number, ranges):
    for range_tuple in ranges:
        if range_tuple[0] <= number <= range_tuple[1]:
            return range_tuple
    return None


def boundary_voltage_calculation(i, MEMS_ID, MEMS_data_array, correction_factor, json_voltage):
    # 定义颜色代码
    # 定义颜色和加粗代码
    RED = '\033[31m'
    BOLD = '\033[1m'
    YELLOW = '\033[92m'
    RESET = '\033[0m'

    cwl_indx = correction_factor[0]
    cwl_warning = correction_factor[1]
    differ_warning = correction_factor[2]

    print('\n################################')
    print(f'-> Info: Anomalies in the boundary!')
    print(f'-> Info: The exception band is {cwl_indx[i]}')

    # 找到 Band 电压对应在 MEMS LUT 表中的索引
    MEMS_LUT_indx = find_closest(MEMS_data_array[:, 1], json_voltage.v1[int(cwl_indx[i] - 1)])
    print(f'-> Info: Voltage index of json file band {cwl_indx[i]} is [{MEMS_LUT_indx}] in MEMS LUT csv file')
    print(
        f'-> Info: Voltage index of Band {cwl_indx[i]} is mapping {MEMS_data_array[MEMS_LUT_indx, 0]} nm in MEMS LUT csv file')

    # 提供 MEMS LUT 表中的电压与波长的单调响应区间
    wave_mems_dis = find_monotonic_intervals(MEMS_data_array[:, 0], 50)
    print(f'-> Info: MEMS ID is [{MEMS_ID}] , it has [{len(wave_mems_dis)}] monotonic interval')
    print(f'-> Info: LUT table range of monotonic interval is {wave_mems_dis}')

    # 判断在 LUT 表的哪个单调区间当中，确定有效单调区间
    range_containing_for_system_band = find_voltage_indx_range_in_LUT(MEMS_LUT_indx, wave_mems_dis)
    print(f'-> Info: Voltage index of json file band {cwl_indx[i]} is in range {range_containing_for_system_band}')

    '''
    将需要修正的系数对齐到 MEMS data LUT 表的波长索引上
    - 20240711: 增加边界条件的判定，如果正好是 abs(value)=5, abs(differ) = 3
    '''
    if abs(differ_warning[i]) == 5:
        if differ_warning[i] < 0:
            # 此时矫正系数定义为 -3
            correct_wave = MEMS_data_array[MEMS_LUT_indx, 0] - 3
        else:
            # 此时矫正系数定义为 3
            correct_wave = MEMS_data_array[MEMS_LUT_indx, 0] + 3
    else:
        correct_wave = MEMS_data_array[MEMS_LUT_indx, 0] + differ_warning[i]
    print(f'-> Info: MEMS wavelength after calibration is {correct_wave} nm')

    # 修正之后的 MEMS 的波长在LUT表中 有效单调区间 中最接近的索引
    MEMS_correct_indx = find_closest(
        MEMS_data_array[min(range_containing_for_system_band): max(range_containing_for_system_band) + 1, 0],
        correct_wave)

    # 确定单调区间内是否存在连续相同的 cwl [连续3个相同作为连续的判断]，找出对应的索引，按照索引最小的赋值即可
    continuous_interval = find_consecutive_triplets(MEMS_data_array[min(range_containing_for_system_band): max(range_containing_for_system_band) + 1, 0])
    continuous_range_indx = find_voltage_indx_range_in_LUT(MEMS_correct_indx, continuous_interval)
    if not continuous_range_indx:
        MEMS_correct_indx = MEMS_correct_indx
    else:
        MEMS_correct_indx = min(continuous_range_indx)


    print(f'-> Info: Indexes corresponding to MEMS wavelengths after calibration is {MEMS_correct_indx}')
    print(f'-> Info: The LUT table corresponds to the wavelength is {MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx), 0]} nm')
    MEMS_correct_voltage = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx), :]

    '''
    # 此时对理想情况下这个计算的 band 的修正参数进行更正，并修正判断代码中的 json 对应的电压
    - 20240711: 增加 MEMS LUT 表中 CWL 间隙过大情况的判定，并进行局部插值计算，此时的条件是在单调区间内 MEMS 的索引在矫正之后并未处于边界状态
    '''
    delta_cal = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx), 0] - MEMS_data_array[MEMS_LUT_indx, 0]
    differ_warning[i] = int(differ_warning[i] - delta_cal)

    if abs(differ_warning[i]) >=5 and min(range_containing_for_system_band) + MEMS_correct_indx + 1 <= max(range_containing_for_system_band):
        print(f'\n{BOLD}{YELLOW}-> Warning: Unable to find the expected index of Band {cwl_indx[i]} correctly! {RESET}')
        print(f'{BOLD}{YELLOW}-> Warning: Attention to the response of the band {cwl_indx[i]} {RESET}!')
        print(f'{BOLD}{YELLOW}-> ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲{RESET}')
        print(f'{BOLD}{YELLOW}-> Info: Voltage calculation using local voltage interpolation!{RESET}')

        # 暂停 3 秒钟并展示进度条
        duration = 1
        progress_bar(duration)

        # 开始电压插值的局部运算,插值采用线性插值法
        mems_local_original = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx), 0]
        meme_local_original_voltage_array = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx), 1 : ]
        mems_local_wanted = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx + 1), 0]
        meme_local_wanted_voltage_array = MEMS_data_array[(min(range_containing_for_system_band) + MEMS_correct_indx)+1,
                                            1: ]

        # 计算差值点数
        voltage_number = int(mems_local_wanted - mems_local_original + 1)
        x_array_original = [mems_local_original, mems_local_wanted]

        # 创建一组新的x值，用于插值计算
        x_array_new = np.linspace(int(x_array_original[0]), int(x_array_original[-1]), num=voltage_number)
        local_voltage_fit = np.zeros([4, voltage_number])

        for count in range(len(meme_local_original_voltage_array)):
            plt.figure(4, figsize=(10, 6))
            plt.title('Local voltage trends')
            plt.plot(x_array_original, [meme_local_original_voltage_array[count], meme_local_wanted_voltage_array[count]], marker='o', color='black', label='Before Voltage Treatment')
            plt.xlabel('Wavelength / nm')
            plt.ylabel('Voltage / V')
            plt.grid(True)

        # 计算局部电压插值数组
        for count in range(len(meme_local_wanted_voltage_array)):
            f_linear = interp1d(x_array_original, [meme_local_original_voltage_array[count], meme_local_wanted_voltage_array[count]])
            local_voltage_fit[count, :] = f_linear(x_array_new)
            plt.plot(x_array_new, local_voltage_fit[count, :], color='blue', linewidth=2,
                     label='After Voltage Treatment Fitting')

        local_correct_indx = find_closest(x_array_new, correct_wave)
        print(f'{BOLD}{YELLOW}-> Info: Indexes corresponding to MEMS wavelengths after calibration is {x_array_new[local_correct_indx]}{RESET}')

        delta_cal = x_array_new[local_correct_indx] - MEMS_data_array[MEMS_LUT_indx, 0]
        differ_warning[i] = int(differ_warning[i] - delta_cal)

        cwl_warning[i] = int(cwl_warning[i] + delta_cal)
        json_voltage.v1[cwl_indx[i] - 1] = local_voltage_fit[0, local_correct_indx]
        json_voltage.v2[cwl_indx[i] - 1] = local_voltage_fit[1, local_correct_indx]
        json_voltage.v3[cwl_indx[i] - 1] = local_voltage_fit[2, local_correct_indx]
        json_voltage.v4[cwl_indx[i] - 1] = local_voltage_fit[3, local_correct_indx]

        for count in range(len(meme_local_original_voltage_array)):
            plt.title('Local voltage trends')
            plt.plot(x_array_new[local_correct_indx], local_voltage_fit[count, local_correct_indx], marker='o', color='red', label='Before Voltage Treatment')
            plt.xlabel('Wavelength / nm')
            plt.ylabel('Voltage / V')
            plt.grid(True)

    else:
        cwl_warning[i] = int(cwl_warning[i] + delta_cal)
        json_voltage.v1[cwl_indx[i] - 1] = MEMS_correct_voltage[1]
        json_voltage.v2[cwl_indx[i] - 1] = MEMS_correct_voltage[2]
        json_voltage.v3[cwl_indx[i] - 1] = MEMS_correct_voltage[3]
        json_voltage.v4[cwl_indx[i] - 1] = MEMS_correct_voltage[4]

    correction_factor = [cwl_indx, cwl_warning, differ_warning]

    return json_voltage, correction_factor

def get_MEMS_voltage_data_array(json_path, model):
    [MEMS_data, MEMS_ID] = get_MEMS_data(json_path, model)
    MEMS_data_array = np.zeros([len(MEMS_data)-1, 5])
    for i in range(len(MEMS_data)-1):
        MEMS_data_array[i, 0] = np.array(MEMS_data[i+1][6])
        for j in range(4):
            MEMS_data_array[i, j+1] = np.array(MEMS_data[i+1][j+1])

    plt.figure(2, figsize=(12, 6))
    # plt.plot(MEMS_data_array[:, 0], MEMS_data_array[:, 1], linewidth=2)
    plt.plot(MEMS_data_array[:, 1], MEMS_data_array[:, 0], linewidth=2)
    # plt.plot(MEMS_data_array[:, 1], linewidth=2)
    plt.xlabel('Voltage / V')
    plt.ylabel('Wavelength / nm')
    plt.grid(True)
    return MEMS_data_array, MEMS_ID


# 判断 json 的电压是否为 1st order 还是 2nd order
def cal_json_voltage_model(arr, threshold=10):
    intervals = []
    start_idx = 0

    for i in range(1, len(arr)):
        if abs(arr[i] - arr[i-1]) > threshold:
            intervals.append((start_idx, i-1))
            start_idx = i

    intervals.append((start_idx, len(arr)-1))  # Add the last interval

    return intervals

def is_monotonic(arr):
    """
    检查一个数组是否为单调区间（要么完全非递减，要么完全非递增）。

    Args:
        arr (list): 输入的数组。

    Returns:
        bool: 如果数组是单调区间，则返回 True，否则返回 False。
    """
    increasing = decreasing = True

    for i in range(1, len(arr)):
        if arr[i] < arr[i - 1]:
            increasing = False
        elif arr[i] > arr[i - 1]:
            decreasing = False

    return increasing or decreasing


def find_monotonic_changes(arr):
    """
    查找数组中单调递增变为单调递减和单调递减变为单调递增的索引。

    Args:
        arr (list): 输入的数组。

    Returns:
        tuple: 包含两个列表，第一个列表为单调递增变为单调递减的索引，第二个列表为单调递减变为单调递增的索引。
    """
    increasing_to_decreasing = []
    decreasing_to_increasing = []

    for i in range(1, len(arr)):
        if arr[i] > arr[i - 1]:
            increasing_to_decreasing.append((i - 1, i))
        elif arr[i] < arr[i - 1]:
            decreasing_to_increasing.append((i - 1, i))

    return increasing_to_decreasing, decreasing_to_increasing

def choose_folder():
    """
    弹出文件夹选择对话框，获取用户选择的文件夹路径。

    Returns:
        str: 用户选择的文件夹路径。
    """
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 获取文件夹路径
    folder_path = filedialog.askdirectory()

    return folder_path

def progress_bar(duration):
    # 定义总步数（例如，100步用于进度条）
    total_steps = 100

    # 计算每一步的时间间隔
    step_interval = duration / total_steps

    # 初始化进度
    progress = 0

    # 打印初始进度条
    print("\033[1;93m[", end="")
    for _ in range(total_steps):
        print(" ", end="")
    print("] 0%\033[0m", end="")

    # 更新进度条
    for step in range(total_steps):
        time.sleep(step_interval)
        progress += 1
        percent_complete = int(progress / total_steps * 100)
        print("\r\033[1;93m[", end="")
        for i in range(total_steps):
            if i < progress:
                print("=", end="")
            else:
                print(" ", end="")
        print(f"] {percent_complete}%\033[0m", end="")

    # 打印进度条完成消息
    print()

def find_consecutive_triplets(arr):
    # 初始化结果列表和临时变量
    result = []
    start = 0

    # 遍历数组
    for i in range(2, len(arr)):
        # 如果当前数字与前两个数字相同
        if arr[i] == arr[i-1] == arr[i-2]:
            # 如果这是新的一组三连数字
            if start != i-2:
                # 将索引范围添加到结果列表
                result.append((i-2, i))
            # 更新起始索引
            start = i-2
        else:
            # 如果当前数字与前两个数字不同，重置起始索引
            start = i

    return result



if __name__ == '__main__':

    selected_folder = choose_folder()
    file_path = selected_folder

    [json_path, json_index] = get_json_folder(file_path)
    [csv_path, model] = get_csv_folder(file_path)

    #################################
    #TODO: 获得LUT表中的 MEMS 电压值
    [MEMS_data_array, MEMS_ID] = get_MEMS_voltage_data_array(json_path[json_index], model)
    MEMS_params = [MEMS_data_array, MEMS_ID]
    ################################

    if not csv_path or not json_path:
        print('-> Info: Something is wrong!!')
        print('-> Info: End!')
    else:
        # 打印找到的文件
        print("-> Info: JSON files ", json_path)
        print("-> Info: CSV files ", csv_path)
        [dataBeSave, correction_factor, exemptions_array] = fitting_voltage(json_path[json_index], csv_path, MEMS_params)
        folder_save_Path = save_fitting_json(file_path, dataBeSave, json_path, json_index, exemptions_array)
        save_xlsx(folder_save_Path, dataBeSave)
        print("-> End!!")



    plt.show()