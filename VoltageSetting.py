################################################################
# 脚本用来实现电压插值.
#
# 输入文件：
# - json文件
# - energyfactor.csv
#
# 输出文件：
# - 插值电压列表.csv
# - 修改后的json
#
# Author: Jerome. Date: 20240626
################################################################

import time
from Function.envi import ENVI
from Function.UNSFunction import *
from datetime import datetime
import json
import csv
import glob
import openpyxl
from datetime import datetime

def get_json_data(json_path):
    # 读取 json 文件内容
    with open(json_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data

def get_csv_data(csv_path):
    # 打开并读取 CSV 文件
    with open(csv_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        # 将内容保存为矩阵
        matrix = [row for row in reader]
        # 打印矩阵内容
        # for row in matrix:
        #     print(row)

    return matrix

class StructVar:
    # 建立存储电压的类
    def __init__(self):
        self.v1 = []
        self.v2 = []
        self.v3 = []
        self.v4 = []


def fitting_voltage(json_path, csv_path):
    energyfactor_index = next((i for i, path in enumerate(csv_path) if 'energyFactor.csv' in path), None)
    # 获取标定的 cwl 矩阵
    csv_data = get_csv_data(csv_path[energyfactor_index])
    cwl_cal = []
    cwl_warning = []
    differ_warning = []
    cwl_indx = []
    for i in range(10):
        data = int(csv_data[i + 1][0])
        cwl_cal.append((data))
        if abs(int(csv_data[i + 1][2])) >= 5:
            print(f'-> Info: band_{i+1} cwl is exceeds grade 5 nm.')
            print(f'-> Info: CWL is {data}, differ is {int(csv_data[i + 1][2])} nm.')
            cwl_warning.append(data)
            differ_warning.append(int(csv_data[i + 1][2]))
            cwl_indx.append(i+1)

    correction_factor = [cwl_indx, cwl_warning, differ_warning]

    ################################

    json_data = get_json_data(json_path)
    band_size = len(json_data['NIR']['MEMS']['Modes'])
    json_voltage_original = []

    json_voltage = StructVar()
    json_voltage_fit = StructVar()

    for i in range(band_size):
        json_voltage_original.append(json_data['NIR']['MEMS']['Modes'][i]['Voltages'])
        json_voltage.v1.append(json_voltage_original[i][0])
        json_voltage.v2.append(json_voltage_original[i][1])
        json_voltage.v3.append(json_voltage_original[i][2])
        json_voltage.v4.append(json_voltage_original[i][3])

    cwl = [713, 736, 759, 782, 805, 828, 851, 874, 897, 920]
    # 计算差值点数
    voltage_number = cwl_cal[-1] - cwl_cal[0] + 1
    x_array_original = cwl_cal

    # 创建一组新的x值，用于插值计算
    x_array_new = np.linspace(cwl_cal[0], cwl_cal[-1], num=voltage_number)
    cwl_array_fit = np.linspace(cwl_cal[0], cwl_cal[-1], num=voltage_number)
    print(f'-> Info: cwl_cal array is {x_array_original}')
    dataBeSave = np.zeros([5, voltage_number])
    dataBeSave[0, :] = cwl_array_fit
    count = 1
    # 循环访问原始电压中的所有变量，进行运算后保存到新的数组中
    for key in json_voltage.__dict__:
        value = getattr(json_voltage, key)
        f_linear = interp1d(x_array_original, value)
        # 进行运算，这里是加10
        setattr(json_voltage_fit, key, f_linear(x_array_new))

        dataBeSave[count, :] = getattr(json_voltage_fit, key)
        count = count + 1

        plt.figure(1, figsize=(10, 6))
        plt.plot(x_array_original, getattr(json_voltage, key), marker='o', color='black')
        plt.plot(x_array_new, getattr(json_voltage_fit, key), color='red', linewidth=2)
        plt.grid(True)

    dataBeSave = dataBeSave.T
    return dataBeSave, correction_factor

def save_scv(file_path, dataBeSave):
    dataHead = ['cwl_cal', 'V1', 'V2', 'V3', 'V4']
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # 将字符串列表写入 Excel 的第一行
    for index, value in enumerate(dataHead, start=1):
        worksheet.cell(row=1, column=index, value=value)

    # 将二维数字列表写入 Excel 从第二行开始的不同行
    for row_index, row_data in enumerate(dataBeSave, start=2):
        for col_index, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    # 生成保存 Excel 文件的完整路径
    excel_file = file_path + r'\voltage_fitting.xlsx'
    # 保存工作簿为 Excel 文件
    try:
        workbook.save(excel_file)
        print(f"-> Info: Data saved to {excel_file}")
    except:
        print('-> Info: Please close csv file and try again.')

def save_fitting_json(file_path, dataBeSave, json_path):
    current_time = datetime.now()
    formatted_time = current_time.strftime("%Y%m%d%H%M")
    # 假设 'original.json' 是原始文件名，'modified.json' 是新文件名
    original_file_name = json_path[0]
    # 提取文件名和扩展名
    file_name_with_extension = os.path.basename(json_path[0])
    # 去掉.json后缀
    file_name_without_extension = os.path.splitext(file_name_with_extension)[0]
    # 创建新json的存放文件夹
    foldername = 'output'
    folder_save_Path = CreateFolder(file_path, foldername)
    new_file_name = folder_save_Path + f'\\{file_name_without_extension}_fitting_{formatted_time}.json'
    print(f'-> Info: new_file_name is {new_file_name}')

    try:
        # 加载原始JSON数据
        json_data = get_json_data(original_file_name)
        for i in range(len(correction_factor[0])):
            print(f'\n-> Info: Now fix the parameters of band {correction_factor[0][i]} ')
            # 原来 json 文件中的电压为
            indx = correction_factor[0][i] - 1
            A = json_data['NIR']['MEMS']['Modes'][indx]['Voltages']
            print(f'-> Info: Original voltage array is {A}')

            # 计算正确的 cwl
            cwl_corrected = correction_factor[1][i] + correction_factor[2][i]
            # 找到 dataBeSave 中正确 cwl 对应的电压
            indexes_of_value = np.where(dataBeSave[:, 0] == cwl_corrected)
            print(f'-> Info: Indx is {indexes_of_value[0]}, corrected cwl is {cwl_corrected}')
            voltage_corrected = [float(dataBeSave[indexes_of_value, 1]), float(dataBeSave[indexes_of_value, 2]), float(dataBeSave[indexes_of_value, 3]), float(dataBeSave[indexes_of_value, 4])]
            print(f'-> Info: Corrected voltage is {voltage_corrected}')

            # 对原来的电压进行修改
            json_data['NIR']['MEMS']['Modes'][indx]['Voltages'] = voltage_corrected

            # 将修改后的数据保存到新的JSON文件
            with open(new_file_name, 'w') as file:
                json.dump(json_data, file, indent=4)

    except FileNotFoundError:
        print(f"文件 '{original_file_name}' 未找到。请确保文件路径正确。")
    except json.JSONDecodeError:
        print(f"文件 '{original_file_name}' 不是有效的JSON格式。")
    except Exception as e:
        print(f"保存文件时出现错误: {e}")


if __name__ == '__main__':
    file_path = r'D:\PythonFunctionLab\Solomon\VoltageSetting\camera242001B8_4th'

    json_path = glob.glob(f"{file_path}/*.json")
    csv_path = glob.glob(f"{file_path}/*.csv")

    # 打印找到的文件
    print("-> Info: JSON files ", json_path)
    print("-> Info: CSV files ", csv_path)


    [dataBeSave, correction_factor] = fitting_voltage(json_path[0], csv_path)
    save_scv(file_path, dataBeSave)
    save_fitting_json(file_path, dataBeSave, json_path)




    # csv_data = get_csv_data(csv_path[0])
    # cwl_cal = []
    # for i in range(10):
    #     cwl_cal.append(csv_data[i+1][0])

    plt.show()